from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta, date
from decimal import Decimal
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Customer, Transaction, AuditLog, UserProfile
from .forms import (
    CustomerForm, TransactionForm,
    ChangePinForm, ChangePasswordForm, ForgotPasswordPinForm,
)


# ─── HELPERS ────────────────────────────────────────────────
def log_action(user, action, target, detail=''):
    AuditLog.objects.create(user=user, action=action, target=target, detail=detail)


def get_or_create_profile(user):
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile


# ─── CUSTOMER LIST ──────────────────────────────────────────
@login_required
def customer_list(request):
    q = request.GET.get('q', '').strip()
    customers = Customer.objects.all()
    if q:
        customers = customers.filter(
            Q(name__icontains=q) | Q(number__icontains=q)
        )

    all_customers   = Customer.objects.all()
    total_customers = all_customers.count()
    total_utang     = sum(c.total_utang() for c in all_customers)
    total_paid      = sum(c.total_paid()  for c in all_customers)
    total_balance   = sum(c.balance()     for c in all_customers)

    paginator = Paginator(customers, 10)
    page      = request.GET.get('page', 1)
    customers = paginator.get_page(page)

    return render(request, 'ledger/customer_list.html', {
        'customers':       customers,
        'q':               q,
        'total_customers': total_customers,
        'total_utang':     total_utang,
        'total_paid':      total_paid,
        'total_balance':   total_balance,
    })


# ─── DASHBOARD ──────────────────────────────────────────────
@login_required
def dashboard(request):
    all_customers = Customer.objects.all()

    total_customers = all_customers.count()
    total_balance   = sum(c.balance()     for c in all_customers)
    total_utang     = sum(c.total_utang() for c in all_customers)
    total_paid      = sum(c.total_paid()  for c in all_customers)

    customers_with_balance = sorted(
        [c for c in all_customers if c.balance() > 0],
        key=lambda c: c.balance(),
        reverse=True
    )[:5]
    top_names    = [c.name for c in customers_with_balance]
    top_balances = [float(c.balance()) for c in customers_with_balance]

    today      = date.today()
    months     = []
    utang_data = []
    paid_data  = []
    for i in range(5, -1, -1):
        d = today.replace(day=1) - timedelta(days=i * 28)
        label = d.strftime('%b %Y')
        months.append(label)
        month_utang = Transaction.objects.filter(
            transaction_type='utang',
            date__year=d.year,
            date__month=d.month
        ).aggregate(t=Sum('amount'))['t'] or 0
        month_paid = Transaction.objects.filter(
            transaction_type='payment',
            date__year=d.year,
            date__month=d.month
        ).aggregate(t=Sum('amount'))['t'] or 0
        utang_data.append(float(month_utang))
        paid_data.append(float(month_paid))

    recent_transactions = Transaction.objects.select_related('customer').all()[:10]
    high_balance = [c for c in all_customers if c.balance() > 1000]

    return render(request, 'ledger/dashboard.html', {
        'total_customers':     total_customers,
        'total_balance':       total_balance,
        'total_utang':         total_utang,
        'total_paid':          total_paid,
        'top_names':           json.dumps(top_names),
        'top_balances':        json.dumps(top_balances),
        'months':              json.dumps(months),
        'utang_data':          json.dumps(utang_data),
        'paid_data':           json.dumps(paid_data),
        'recent_transactions': recent_transactions,
        'high_balance':        high_balance,
    })


# ─── CUSTOMER ADD ────────────────────────────────────────────
@login_required
def customer_add(request):
    form = CustomerForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        customer = form.save()
        log_action(request.user, 'create', f'Customer: {customer.name}')
        messages.success(request, "Customer saved successfully!")
        return redirect('customer_list')
    return render(request, 'ledger/customer_form.html', {
        'form':  form,
        'title': 'Add Customer',
        'btn':   'Save Customer',
    })


# ─── CUSTOMER EDIT ───────────────────────────────────────────
@login_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    form     = CustomerForm(request.POST or None, instance=customer)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_action(request.user, 'edit', f'Customer: {customer.name}')
        messages.success(request, "Customer updated!")
        return redirect('customer_detail', pk=pk)
    return render(request, 'ledger/customer_form.html', {
        'form':     form,
        'title':    f'Edit — {customer.name}',
        'btn':      'Update Customer',
        'customer': customer,
    })


# ─── CUSTOMER DELETE ─────────────────────────────────────────
@login_required
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        name = customer.name
        log_action(request.user, 'delete', f'Customer: {name}', f'Had balance: {customer.balance()}')
        customer.delete()
        messages.success(request, f"Customer '{name}' deleted successfully!")
        return redirect('customer_list')
    return render(request, 'ledger/customer_confirm_delete.html', {'customer': customer})


# ─── CUSTOMER DETAIL ─────────────────────────────────────────
@login_required
def customer_detail(request, pk):
    customer     = get_object_or_404(Customer, pk=pk)
    transactions = customer.transactions.all()

    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    if date_from:
        transactions = transactions.filter(date__date__gte=date_from)
    if date_to:
        transactions = transactions.filter(date__date__lte=date_to)

    paginator    = Paginator(transactions, 15)
    page         = request.GET.get('page', 1)
    transactions = paginator.get_page(page)

    return render(request, 'ledger/customer_detail.html', {
        'customer':     customer,
        'transactions': transactions,
        'date_from':    date_from,
        'date_to':      date_to,
    })


# ─── CUSTOMER PRINT ──────────────────────────────────────────
@login_required
def customer_print(request, pk):
    customer     = get_object_or_404(Customer, pk=pk)
    transactions = customer.transactions.all()
    return render(request, 'ledger/customer_print.html', {
        'customer':     customer,
        'transactions': transactions,
        'printed_at':   timezone.now(),
    })


# ─── TRANSACTION ADD ─────────────────────────────────────────
@login_required
def transaction_add(request, customer_pk):
    customer = get_object_or_404(Customer, pk=customer_pk)
    form     = TransactionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        t          = form.save(commit=False)
        t.customer = customer
        t.save()
        log_action(request.user, 'create',
                   f'Transaction for {customer.name}',
                   f'{t.transaction_type} ₱{t.amount}')
        messages.success(request, "Transaction saved!")
        return redirect('customer_detail', pk=customer_pk)
    return render(request, 'ledger/transaction_form.html', {
        'form':     form,
        'customer': customer,
        'title':    'Add Transaction',
        'btn':      'Save Transaction',
    })


# ─── TRANSACTION EDIT ────────────────────────────────────────
@login_required
def transaction_edit(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    form        = TransactionForm(request.POST or None, instance=transaction)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_action(request.user, 'edit',
                   f'Transaction for {transaction.customer.name}',
                   f'{transaction.transaction_type} ₱{transaction.amount}')
        messages.success(request, "Transaction updated!")
        return redirect('customer_detail', pk=transaction.customer.pk)
    return render(request, 'ledger/transaction_form.html', {
        'form':     form,
        'customer': transaction.customer,
        'title':    'Edit Transaction',
        'btn':      'Update Transaction',
    })


# ─── TRANSACTION DELETE ──────────────────────────────────────
@login_required
def transaction_delete(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    customer_pk = transaction.customer.pk
    if request.method == 'POST':
        log_action(request.user, 'delete',
                   f'Transaction for {transaction.customer.name}',
                   f'{transaction.transaction_type} ₱{transaction.amount}')
        transaction.delete()
        messages.success(request, "Transaction deleted!")
        return redirect('customer_detail', pk=customer_pk)
    return render(request, 'ledger/transaction_confirm_delete.html', {'transaction': transaction})


# ─── AUDIT LOG ───────────────────────────────────────────────
@login_required
def audit_log_view(request):
    logs      = AuditLog.objects.select_related('user').all()
    paginator = Paginator(logs, 20)
    page      = request.GET.get('page', 1)
    logs      = paginator.get_page(page)
    return render(request, 'ledger/audit_log.html', {'logs': logs})


# ─── CHANGE PIN ──────────────────────────────────────────────
@login_required
def change_pin(request):
    profile = get_or_create_profile(request.user)
    form    = ChangePinForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        current = form.cleaned_data.get('current_pin', '')
        new_pin = form.cleaned_data['new_pin']
        hint    = form.cleaned_data.get('pin_hint', '')

        if profile.pin and current != profile.pin:
            messages.error(request, 'Current PIN is incorrect.')
            return render(request, 'ledger/change_pin.html', {'form': form, 'profile': profile})

        profile.pin      = new_pin
        profile.pin_hint = hint
        profile.save()
        log_action(request.user, 'edit', 'PIN', 'PIN changed')
        messages.success(request, 'PIN updated successfully!')
        return redirect('dashboard')

    return render(request, 'ledger/change_pin.html', {'form': form, 'profile': profile})


# ─── CHANGE PASSWORD ─────────────────────────────────────────
@login_required
def change_password(request):
    form = ChangePasswordForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        current = form.cleaned_data['current_password']
        new_pw  = form.cleaned_data['new_password']

        user = authenticate(username=request.user.username, password=current)
        if user is None:
            messages.error(request, 'Current password is incorrect.')
            return render(request, 'ledger/change_password.html', {'form': form})

        user.set_password(new_pw)
        user.save()
        update_session_auth_hash(request, user)
        log_action(request.user, 'edit', 'Password', 'Password changed')
        messages.success(request, 'Password updated successfully!')
        return redirect('dashboard')

    return render(request, 'ledger/change_password.html', {'form': form})


# ─── FORGOT PASSWORD (via PIN) ───────────────────────────────
def forgot_password(request):
    form = ForgotPasswordPinForm(request.POST or None)
    hint = None

    username_typed = request.GET.get('username', '').strip() or request.POST.get('username', '').strip()
    if username_typed:
        try:
            u       = User.objects.get(username=username_typed)
            profile = get_or_create_profile(u)
            hint    = profile.pin_hint or None
        except User.DoesNotExist:
            pass

    if request.method == 'POST' and form.is_valid():
        username = form.cleaned_data['username']
        pin      = form.cleaned_data['pin']
        new_pw   = form.cleaned_data['new_password']

        try:
            user    = User.objects.get(username=username)
            profile = get_or_create_profile(user)
        except User.DoesNotExist:
            messages.error(request, 'Username not found.')
            return render(request, 'ledger/forgot_password.html', {'form': form, 'hint': hint})

        if not profile.pin:
            messages.error(request, 'No PIN set for this account. Contact your admin.')
            return render(request, 'ledger/forgot_password.html', {'form': form, 'hint': hint})

        if pin != profile.pin:
            messages.error(request, 'Incorrect PIN.')
            return render(request, 'ledger/forgot_password.html', {'form': form, 'hint': hint})

        user.set_password(new_pw)
        user.save()
        messages.success(request, 'Password reset! You may now log in.')
        return redirect('login')

    return render(request, 'ledger/forgot_password.html', {'form': form, 'hint': hint})


# ─── EXPORT: ALL CUSTOMERS EXCEL ─────────────────────────────
@login_required
def export_customers_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")

    headers = ['#', 'Name', 'Phone Number', 'Total Utang', 'Total Paid', 'Balance', 'Date Added']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, customer in enumerate(Customer.objects.all(), 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=customer.name)
        ws.cell(row=row, column=3, value=customer.number or '—')
        ws.cell(row=row, column=4, value=float(customer.total_utang()))
        ws.cell(row=row, column=5, value=float(customer.total_paid()))
        ws.cell(row=row, column=6, value=float(customer.balance()))
        ws.cell(row=row, column=7, value=customer.created_at.strftime('%Y-%m-%d'))

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
    wb.save(response)
    return response


# ─── EXPORT: CUSTOMER TRANSACTIONS EXCEL ─────────────────────
@login_required
def export_customer_transactions_excel(request, pk):
    customer     = get_object_or_404(Customer, pk=pk)
    transactions = customer.transactions.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = customer.name[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")

    headers = ['#', 'Type', 'Amount', 'Note', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=t.get_transaction_type_display())
        ws.cell(row=row, column=3, value=float(t.amount))
        ws.cell(row=row, column=4, value=t.note or '—')
        ws.cell(row=row, column=5, value=t.date.strftime('%Y-%m-%d %H:%M'))

    ws.append([])
    ws.append(['', 'Total Utang', float(customer.total_utang())])
    ws.append(['', 'Total Paid',  float(customer.total_paid())])
    ws.append(['', 'Balance',     float(customer.balance())])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{customer.name}_transactions.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── EXPORT: BACKUP JSON ─────────────────────────────────────
@login_required
def export_backup_json(request):
    data = {
        'customers':    list(Customer.objects.all().values()),
        'transactions': list(Transaction.objects.all().values()),
    }
    response = HttpResponse(
        content=json.dumps(data, indent=2, default=str),
        content_type='application/json'
    )
    response['Content-Disposition'] = 'attachment; filename="utang_backup.json"'
    return response